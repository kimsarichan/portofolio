import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import chisquare
import ast
import tokenize
import json
import re
from openpyxl import Workbook
from StringIO import StringIO
import openpyxl

harga={
  "goods": [
    {
      "Id": "0",
      "Flavor": "Chocolate",
      "Food": "Cake",
      "Price": "8.95",
      "Type": "Food"
    },
    {
      "Id": "1",
      "Flavor": "Lemon",
      "Food": "Cake",
      "Price": "8.95",
      "Type": "Food"
    },
    {
      "Id": "2",
      "Flavor": "Casino",
      "Food": "Cake",
      "Price": "15.95",
      "Type": "Food"
    },
    {
      "Id": "3",
      "Flavor": "Opera",
      "Food": "Cake",
      "Price": "15.95",
      "Type": "Food"
    },
    {
      "Id": "4",
      "Flavor": "Strawberry",
      "Food": "Cake",
      "Price": "11.95",
      "Type": "Food"
    },
    {
      "Id": "5",
      "Flavor": "Truffle",
      "Food": "Cake",
      "Price": "15.95",
      "Type": "Food"
    },
    {
      "Id": "6",
      "Flavor": "Chocolate",
      "Food": "Eclair",
      "Price": "3.25",
      "Type": "Food"
    },
    {
      "Id": "7",
      "Flavor": "Coffee",
      "Food": "Eclair",
      "Price": "3.5",
      "Type": "Food"
    },
    {
      "Id": "8",
      "Flavor": "Vanilla",
      "Food": "Eclair",
      "Price": "3.25",
      "Type": "Food"
    },
    {
      "Id": "9",
      "Flavor": "Napoleon",
      "Food": "Cake",
      "Price": "13.49",
      "Type": "Food"
    },
    {
      "Id": "10",
      "Flavor": "Almond",
      "Food": "Tart",
      "Price": "3.75",
      "Type": "Food"
    },
    {
      "Id": "11",
      "Flavor": "Apple",
      "Food": "Pie",
      "Price": "5.25",
      "Type": "Food"
    },
    {
      "Id": "12",
      "Flavor": "Apple",
      "Food": "Tart",
      "Price": "3.25",
      "Type": "Food"
    },
    {
      "Id": "13",
      "Flavor": "Apricot",
      "Food": "Tart",
      "Price": "3.25",
      "Type": "Food"
    },
    {
      "Id": "14",
      "Flavor": "Berry",
      "Food": "Tart",
      "Price": "3.25",
      "Type": "Food"
    },
    {
      "Id": "15",
      "Flavor": "Blackberry",
      "Food": "Tart",
      "Price": "3.25",
      "Type": "Food"
    },
    {
      "Id": "16",
      "Flavor": "Blueberry",
      "Food": "Tart",
      "Price": "3.25",
      "Type": "Food"
    },
    {
      "Id": "17",
      "Flavor": "Chocolate",
      "Food": "Tart",
      "Price": "3.75",
      "Type": "Food"
    },
    {
      "Id": "18",
      "Flavor": "Cherry",
      "Food": "Tart",
      "Price": "3.25",
      "Type": "Food"
    },
    {
      "Id": "19",
      "Flavor": "Lemon",
      "Food": "Tart",
      "Price": "3.25",
      "Type": "Food"
    },
    {
      "Id": "20",
      "Flavor": "Pecan",
      "Food": "Tart",
      "Price": "3.75",
      "Type": "Food"
    },
    {
      "Id": "21",
      "Flavor": "Ganache",
      "Food": "Cookie",
      "Price": "1.15",
      "Type": "Food"
    },
    {
      "Id": "22",
      "Flavor": "Gongolais",
      "Food": "Cookie",
      "Price": "1.15",
      "Type": "Food"
    },
    {
      "Id": "23",
      "Flavor": "Raspberry",
      "Food": "Cookie",
      "Price": "1.09",
      "Type": "Food"
    },
    {
      "Id": "24",
      "Flavor": "Lemon",
      "Food": "Cookie",
      "Price": "0.79",
      "Type": "Food"
    },
    {
      "Id": "25",
      "Flavor": "Chocolate",
      "Food": "Meringue",
      "Price": "1.25",
      "Type": "Food"
    },
    {
      "Id": "26",
      "Flavor": "Vanilla",
      "Food": "Meringue",
      "Price": "1.15",
      "Type": "Food"
    },
    {
      "Id": "27",
      "Flavor": "Marzipan",
      "Food": "Cookie",
      "Price": "1.25",
      "Type": "Food"
    },
    {
      "Id": "28",
      "Flavor": "Tuile",
      "Food": "Cookie",
      "Price": "1.25",
      "Type": "Food"
    },
    {
      "Id": "29",
      "Flavor": "Walnut",
      "Food": "Cookie",
      "Price": "0.79",
      "Type": "Food"
    },
    {
      "Id": "30",
      "Flavor": "Almond",
      "Food": "Croissant",
      "Price": "1.45",
      "Type": "Food"
    },
    {
      "Id": "31",
      "Flavor": "Apple",
      "Food": "Croissant",
      "Price": "1.45",
      "Type": "Food"
    },
    {
      "Id": "32",
      "Flavor": "Apricot",
      "Food": "Croissant",
      "Price": "1.45",
      "Type": "Food"
    },
    {
      "Id": "33",
      "Flavor": "Cheese",
      "Food": "Croissant",
      "Price": "1.75",
      "Type": "Food"
    },
    {
      "Id": "34",
      "Flavor": "Chocolate",
      "Food": "Croissant",
      "Price": "1.75",
      "Type": "Food"
    },
    {
      "Id": "35",
      "Flavor": "Apricot",
      "Food": "Danish",
      "Price": "1.15",
      "Type": "Food"
    },
    {
      "Id": "36",
      "Flavor": "Apple",
      "Food": "Danish",
      "Price": "1.15",
      "Type": "Food"
    },
    {
      "Id": "37",
      "Flavor": "Almond",
      "Food": "Twist",
      "Price": "1.15",
      "Type": "Food"
    },
    {
      "Id": "38",
      "Flavor": "Almond",
      "Food": "Bear Claw",
      "Price": "1.95",
      "Type": "Food"
    },
    {
      "Id": "39",
      "Flavor": "Blueberry",
      "Food": "Danish",
      "Price": "1.15",
      "Type": "Food"
    },
    {
      "Id": "40",
      "Flavor": "Lemon",
      "Food": "Lemonade",
      "Price": "3.25",
      "Type": "Drink"
    },
    {
      "Id": "41",
      "Flavor": "Raspberry",
      "Food": "Lemonade",
      "Price": "3.25",
      "Type": "Drink"
    },
    {
      "Id": "42",
      "Flavor": "Orange",
      "Food": "Juice",
      "Price": "2.75",
      "Type": "Drink"
    },
    {
      "Id": "43",
      "Flavor": "Green",
      "Food": "Tea",
      "Price": "1.85",
      "Type": "Drink"
    },
    {
      "Id": "44",
      "Flavor": "Bottled",
      "Food": "Water",
      "Price": "1.8",
      "Type": "Drink"
    },
    {
      "Id": "45",
      "Flavor": "Hot",
      "Food": "Coffee",
      "Price": "2.15",
      "Type": "Drink"
    },
    {
      "Id": "46",
      "Flavor": "Chocolate",
      "Food": "Coffee",
      "Price": "2.45",
      "Type": "Drink"
    },
    {
      "Id": "47",
      "Flavor": "Vanilla",
      "Food": "Frappuccino",
      "Price": "3.85",
      "Type": "Drink"
    },
    {
      "Id": "48",
      "Flavor": "Cherry",
      "Food": "Soda",
      "Price": "1.29",
      "Type": "Drink"
    },
    {
      "Id": "49",
      "Flavor": "Single",
      "Food": "Espresso",
      "Price": "1.85",
      "Type": "Drink');"
    }
  ]
}
distribusi =[{"name":"toko1","item":[
                {"dist":"duniform","a":15,"b":51},
                {"dist":"negbinomial","n":17,"p":0.3547},
                {"dist":"duniform","a":15,"b":42},
                {"dist":"duniform","a":14,"b":48},
                {"dist":"duniform","a":14,"b":58},
                {"dist":"duniform","a":22,"b":46},
                {"dist":"duniform","a":5,"b":31},
                {"dist":"duniform","a":22,"b":61},
                {"dist":"negbinomial","n":6,"p":0.27286},
                {"dist":"duniform","a":16,"b":46},
                {"dist":"duniform","a":4,"b":29},
                {"dist":"negbinomial","n":21,"p":0.41261},
                {"dist":"duniform","a":13,"b":39},
                {"dist":"duniform","a":2,"b":34},
                {"dist":"duniform","a":14,"b":48},
                {"dist":"duniform","a":14,"b":46},
                {"dist":"negbinomial","n":9,"p":0.23118},
                {"dist":"duniform","a":12,"b":45},
                {"dist":"negbinomial","n":24,"p":0.39311},
                {"dist":"negbinomial","n":16,"p":0.35136},
                {"dist":"duniform","a":4,"b":31},
                {"dist":"duniform","a":4,"b":28},
                {"dist":"duniform","a":19,"b":56},
                {"dist":"duniform","a":10,"b":42},
                {"dist":"duniform","a":5,"b":46},
                {"dist":"duniform","a":2,"b":28},
                {"dist":"duniform","a":7,"b":30},
                {"dist":"duniform","a":22,"b":56},
                {"dist":"negbinomial","n":26,"p":0.36764},
                {"dist":"duniform","a":13,"b":42},
                {"dist":"negbinomial","n":7,"p":0.30519},
                {"dist":"negbinomial","n":9,"p":0.24367},
                {"dist":"duniform","a":12,"b":48},
                {"dist":"duniform","a":14,"b":54},
                {"dist":"duniform","a":6,"b":30},
                {"dist":"negbinomial","n":19,"p":0.3572},
                {"dist":"duniform","a":6,"b":39},
                {"dist":"poisson","lambda":26.241},
                {"dist":"duniform","a":2,"b":28},
                {"dist":"duniform","a":2,"b":34},
                {"dist":"negbinomial","n":17,"p":0.38318},
                {"dist":"duniform","a":7,"b":47},
                {"dist":"duniform","a":17,"b":60},
                {"dist":"duniform","a":10,"b":38},
                {"dist":"negbinomial","n":8,"p":0.23498},
                {"dist":"negbinomial","n":36,"p":0.47982},
                {"dist":"negbinomial","n":12,"p":0.26346},
                {"dist":"duniform","a":15,"b":48},
                {"dist":"duniform","a":11,"b":36},
                {"dist":"duniform","a":11,"b":38}]},
             {"name":"toko2","item":[
                {"dist":"duniform","a":18,"b":56},
                {"dist":"duniform","a":17,"b":46},
                {"dist":"poisson","lambda":32.862},
                {"dist":"poisson","lambda":33.379},
                {"dist":"negbinomial","n":10,"p":0.21808},
                {"dist":"duniform","a":12,"b":53},
                {"dist":"duniform","a":3,"b":27},
                {"dist":"negbinomial","n":19,"p":0.29355},
                {"dist":"duniform","a":6,"b":28},
                {"dist":"duniform","a":17,"b":48},
                {"dist":"duniform","a":3,"b":25},
                {"dist":"negbinomial","n":18,"p":0.36307},
                {"dist":"negbinomial","n":9,"p":0.26712},
                {"dist":"duniform","a":5,"b":31},
                {"dist":"negbinomial","n":33,"p":0.50255},
                {"dist":"negbinomial","n":11,"p":0.28256},
                {"dist":"negbinomial","n":23,"p":0.41205},
                {"dist":"negbinomial","n":10,"p":0.28469},
                {"dist":"negbinomial","n":14,"p":0.29825},
                {"dist":"duniform","a":11,"b":50},
                {"dist":"poisson","lambda":15.586},
                {"dist":"duniform","a":3,"b":32},
                {"dist":"duniform","a":23,"b":59},
                {"dist":"duniform","a":11,"b":47},
                {"dist":"negbinomial","n":39,"p":0.56589},
                {"dist":"negbinomial","n":7,"p":0.29436},
                {"dist":"duniform","a":6,"b":30},
                {"dist":"duniform","a":16,"b":52},
                {"dist":"negbinomial","n":15,"p":0.25928},
                {"dist":"duniform","a":10,"b":37},
                {"dist":"negbinomial","n":9,"p":0.32879},
                {"dist":"duniform","a":11,"b":45},
                {"dist":"duniform","a":18,"b":45},
                {"dist":"duniform","a":19,"b":44},
                {"dist":"duniform","a":4,"b":35},
                {"dist":"poisson","lambda":34.138},
                {"dist":"duniform","a":12,"b":42},
                {"dist":"duniform","a":16,"b":45},
                {"dist":"duniform","a":2,"b":29},
                {"dist":"duniform","a":1,"b":29},
                {"dist":"duniform","a":8,"b":45},
                {"dist":"duniform","a":9,"b":46},
                {"dist":"poisson","lambda":33.31},
                {"dist":"duniform","a":11,"b":38},
                {"dist":"duniform","a":13,"b":42},
                {"dist":"duniform","a":18,"b":66},
                {"dist":"duniform","a":17,"b":50},
                {"dist":"duniform","a":15,"b":42},
                {"dist":"poisson","lambda":24.793},
                {"dist":"duniform","a":11,"b":43}]},
             ]
def createexcel(data,harga):
    wb=Workbook()
    #for i in range(1,21):
    for k in range(0,2):
        n=1
        toko="toko"+str(k)
        ws=wb.create_sheet(toko)
        for i,j in zip(data[k]["item"],harga["goods"]):
            ws.cell(row=n, column=1).value = n
            ws.cell(row=n, column=2).value = j["Food"]
            ws.cell(row=n, column=3).value = i["dist"]
            if(i["dist"]=="duniform"):
                ws.cell(row=n, column=4).value = "a="+str(i["a"])+", b="+str(i["b"])
            if(i["dist"]=="poisson"):
                ws.cell(row=n, column=4).value = "lambda="+str(i["lambda"])
            if(i["dist"]=="negbinomial"):
                ws.cell(row=n, column=4).value = "n="+str(i["n"])+", p="+str(i["p"])
            n+=1
    nama="hasil-distribusi.xlsx"
    wb.save(nama)
def mse(actual,predicted):
    mse_hasil = 0
    for i, j in zip(actual,predicted):
        mse_hasil =mse_hasil+ (float(i)-float(j))**2
    return mse_hasil
#createexcel(distribusi,harga)


skenario_toko_1 = [{"1":33,"2":31,"3":28,"4":30,"5":36,"6":34,"7":17,"8":41,"9":16,"10":31,"11":16,"12":30,"13":26,"14":17,"15":31,"16":30,"17":31,"18":28,"19":37,"20":30,"21":17,"22":16,"23":37,"24":26,"25":25,"26":14,"27":18,"28":39,"29":46,"30":27,"31":17,"32":28,"33":30,"34":33,"35":17,"36":34,"37":22,"38":26,"39":14,"40":17,"41":27,"42":27,"43":38,"44":23,"45":28,"46":39,"47":33,"48":31,"49":23,"50":24},
{"1":54,"2":55,"3":45,"4":56,"5":71,"6":48,"7":39,"8":63,"9":34,"10":48,"11":35,"12":49,"13":41,"14":39,"15":52,"16":46,"17":74,"18":47,"19":59,"20":50,"21":39,"22":30,"23":59,"24":45,"25":49,"26":34,"27":33,"28":57,"29":69,"30":47,"31":36,"32":52,"33":63,"34":68,"35":34,"36":61,"37":43,"38":49,"39":34,"40":40,"41":48,"42":60,"43":66,"44":43,"45":56,"46":58,"47":62,"48":53,"49":40,"50":41},
{"1":43,"2":43,"3":36,"4":43,"5":53,"6":41,"7":28,"8":52,"9":25,"10":39,"11":25,"12":39,"13":33,"14":28,"15":41,"16":38,"17":52,"18":37,"19":48,"20":40,"21":28,"22":23,"23":48,"24":35,"25":37,"26":24,"27":25,"28":48,"29":57,"30":37,"31":26,"32":40,"33":46,"34":50,"35":25,"36":47,"37":32,"38":37,"39":24,"40":28,"41":37,"42":43,"43":52,"44":33,"45":42,"46":48,"47":47,"48":42,"49":31,"50":32}];

skenario_toko_2 = [{"1":37,"2":31,"3":32,"4":33,"5":38,"6":32,"7":15,"8":45,"9":17,"10":32,"11":14,"12":32,"13":25,"14":17,"15":33,"16":29,"17":33,"18":25,"19":34,"20":30,"21":15,"22":17,"23":40,"24":28,"25":30,"26":17,"27":18,"28":34,"29":43,"30":23,"31":18,"32":28,"33":31,"34":31,"35":19,"36":34,"37":27,"38":30,"39":15,"40":14,"41":26,"42":27,"43":33,"44":24,"45":27,"46":41,"47":33,"48":28,"49":24,"50":27},
{"1":59,"2":52,"3":62,"4":64,"5":67,"6":63,"7":28,"8":77,"9":30,"10":48,"11":29,"12":59,"13":45,"14":37,"15":48,"16":48,"17":58,"18":49,"19":58,"20":54,"21":30,"22":34,"23":63,"24":56,"25":49,"26":33,"27":33,"28":56,"29":75,"30":45,"31":38,"32":48,"33":45,"34":51,"35":39,"36":64,"37":44,"38":49,"39":32,"40":36,"41":54,"42":50,"43":54,"44":53,"45":42,"46":73,"47":53,"48":43,"49":42,"50":49},
{"1":48,"2":41,"3":47,"4":48,"5":52,"6":47,"7":21,"8":61,"9":23,"10":40,"11":21,"12":45,"13":35,"14":27,"15":40,"16":38,"17":45,"18":37,"19":46,"20":42,"21":22,"22":25,"23":51,"24":42,"25":39,"26":25,"27":25,"28":45,"29":59,"30":34,"31":28,"32":38,"33":38,"34":41,"35":29,"36":49,"37":35,"38":39,"39":23,"40":25,"41":40,"42":38,"43":43,"44":38,"45":34,"46":57,"47":43,"48":35,"49":33,"50":38}];

keuntungan=0.3
hargabeli=0.7
#print len(distribusi[0]["item"])
#print len(distribusi[1]["item"])
def generatestok(toko,no_skenario):
    hasil=[]
    for i in range(1,51):
        a=[]
        for j in range(29):
            if(toko=="toko1"):
                a.append(skenario_toko_1[no_skenario-1][str(i)])
            else:
                a.append(skenario_toko_2[no_skenario-1][str(i)])
        hasil.append(a)
    return hasil

def distribusi_func(data,toko):
    hasil=[]
    #print len(data[toko-1]["item"])
    for i in (data[toko-1]["item"]):
        if(i["dist"]=="poisson"):
            a=np.random.poisson(lam=i["lambda"],size=29)
            hasil.append(a)
        elif(i["dist"]=="negbinomial"):
            a=np.random.negative_binomial(n=i["n"],p=i["p"],size=29)
            hasil.append(a)
        elif(i["dist"]=="duniform"):
            a=np.random.randint(low=i["a"],high=i["b"],size=29)
            hasil.append(a)
    return hasil

def getData():
    def split_coba(line):
        while line[0]!='{':
            line = line.replace(line[0],"",1)
        return line

    f = open('mosi.txt')
    a=[]
    di=[]
    v={}
    st='Toko ke-'
    i=0
    tgl = 1
    for line in f.readlines():
        if(st in line):
            v[i]=di
            di=[]
            i+=1
            a=[]
            continue
        data = split_coba(line).replace("{","").replace("}","").split(",")
        abc = []
        for ab in data:
            abc.append(ab.replace("\n","").split(": ")[1])
        di.append(abc)
        if(i>2):
            break

    def get_barang(toko,barang,data):
        array_barang=[]
        for k in data[toko]:
            array_barang.append(k[barang])
        return array_barang

    def write_excel(data):
        wb=[]
        for i in range(1,3):
            name="toko"+str(i)
            ws=[]
            for j in range(0,50):
                wj=[]
                bar= get_barang(i,j,data)
                r=1
                for statN in bar:
                    wj.append(int(statN))
                ws.append(wj)
            wb.append(ws)
        return wb
    return write_excel(v)

#inisialisasi
hasiltoko = [[],[]]
hasiltoko[0]=(distribusi_func(distribusi,1))
hasiltoko[1]=(distribusi_func(distribusi,2))

data = getData()

for i in range(2):
    for j in range(50):
        print chisquare(f_obs=hasiltoko[i][j],f_exp=data[i][j])

stoktoko = [[[],[],[]],[[],[],[]]]

stoktoko[0][0] = generatestok("toko1",1)
stoktoko[0][1] = generatestok("toko1",2)
stoktoko[0][2] = generatestok("toko1",3)

stoktoko[1][0] = generatestok("toko2",1)
stoktoko[1][1] = generatestok("toko2",2)
stoktoko[1][2] = generatestok("toko2",3)

def sisa(stok, demand ):
    hasil=[]
    for i in demand:
        item=[]
        for j in i :
            sisa = stok-distribusi
            item.append(sisa)
        hasil.append(item)
    return hasil

def total_keuntungan_item_day(demand,stok,harga):
    terjual=0
    if(demand<stok):
        terjual=demand
    elif(demand>=stok):
        terjual=stok
    hasil= ((keuntungan*harga)*terjual)-((hargabeli*harga)*(stok-terjual))
    return hasil

def total_opportunity_item_day(demand,stok,harga):
    if (stok-demand)<0:
        return (keuntungan*harga)*(abs(stok-demand))
    else:
        return 0

def total_keuntungan_all(demand, harga, stok):
    untung=[]
    for i in harga["goods"]:
        hasil =[]
        id_item =int(i["Id"])
        harga_fix=float(i["Price"])
        k=0
        for j in range(len(demand[id_item])):
            untung_item = total_keuntungan_item_day(demand[id_item][j],stok[id_item][j],harga_fix)
            hasil.append(untung_item)
        untung.append(hasil)
    return untung

def total_opportunity_all(demand, harga, stok):
    untung=[]
    for i in harga["goods"]:
        hasil =[]
        id_item =int(i["Id"])
        harga_fix=float(i["Price"])
        k=0
        for j in range(len(demand[id_item])):
            untung_item = total_opportunity_item_day(demand[id_item][j],stok[id_item][j],harga_fix)
            hasil.append(untung_item)
        untung.append(hasil)
    return untung

#keuntungan
keuntungan_t1_s1 = total_keuntungan_all(hasiltoko[0],harga,stoktoko[0][0])
keuntungan_t1_s2 = total_keuntungan_all(hasiltoko[0],harga,stoktoko[0][1])
keuntungan_t1_s3 = total_keuntungan_all(hasiltoko[0],harga,stoktoko[0][2])

keuntungan_t2_s1 = total_keuntungan_all(hasiltoko[1],harga,stoktoko[1][0])
keuntungan_t2_s2 = total_keuntungan_all(hasiltoko[1],harga,stoktoko[1][1])
keuntungan_t2_s3 = total_keuntungan_all(hasiltoko[1],harga,stoktoko[1][2])

keuntungan_toko = [[keuntungan_t1_s1,keuntungan_t1_s2,keuntungan_t1_s3],
                   [keuntungan_t2_s1,keuntungan_t2_s2,keuntungan_t2_s3]]

#oportunity
oportunity_t1_s1 = total_opportunity_all(hasiltoko[0],harga,stoktoko[0][0])
oportunity_t1_s2 = total_opportunity_all(hasiltoko[0],harga,stoktoko[0][1])
oportunity_t1_s3 = total_opportunity_all(hasiltoko[0],harga,stoktoko[0][2])

oportunity_t2_s1 = total_opportunity_all(hasiltoko[1],harga,stoktoko[1][0])
oportunity_t2_s2 = total_opportunity_all(hasiltoko[1],harga,stoktoko[1][1])
oportunity_t2_s3 = total_opportunity_all(hasiltoko[1],harga,stoktoko[1][2])

oportunity_toko = [[oportunity_t1_s1,oportunity_t1_s2,oportunity_t1_s3],
                   [oportunity_t2_s1,oportunity_t2_s2,oportunity_t2_s3]]

def grafik():
    for toko in range(2):
        for skenario in range(3):
            plt.suptitle("Toko "+str(toko)+", Skenario "+str(skenario)+
                         ", Keuntungan="+str(sum([sum(i) for i in keuntungan_toko[toko][skenario]]))+
                         ", Oportunity Cost="+str(sum([sum(i) for i in oportunity_toko[toko][skenario]])))
            for i,j in zip(hasiltoko[toko],range(1,51)):
                plt.subplot(10,5,j)
                title = "p=" + str(j) + ", u=" + str(sum(keuntungan_toko[toko][skenario][j-1]))+\
                        ", o="+str(sum(oportunity_toko[toko][skenario][j-1]))
                plt.title(title,fontsize=10)
                days = [_ for _ in range(29)]
                plt.plot(days,i,'b-',days,stoktoko[toko][skenario][j-1],'r-')
                plt.yticks(fontsize=6)
                plt.xticks(fontsize=6)
            plt.show()

grafik()
